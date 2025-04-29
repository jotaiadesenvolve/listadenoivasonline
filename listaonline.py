import streamlit as st
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from PIL import Image

# Configuração da página
st.set_page_config(page_title="🎀 Lista de Presentes de Casamento 🎀", layout="centered")

# Senha para edição
SENHA_CORRETA = "casamento123"

# Funções de carregamento e salvamento
def carregar_presentes():
    if os.path.exists("presentes.json"):
        with open("presentes.json", "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def salvar_presentes(presentes):
    with open("presentes.json", "w", encoding="utf-8") as f:
        json.dump(presentes, f, indent=4, ensure_ascii=False)

# Função para criar pasta de imagens
def criar_pasta_imagens():
    if not os.path.exists("fotos_presentes"):
        os.makedirs("fotos_presentes")

# Inicializações
presentes = carregar_presentes()
criar_pasta_imagens()

# Estilo de boas-vindas
st.markdown("<h1 style='text-align: center; color: #DAA520;'>💍 Lista de Presentes de Casamento 💍</h1>", unsafe_allow_html=True)
st.markdown("<h4 style='text-align: center;'>Escolha um presente para os noivos com carinho! 🎁</h4>", unsafe_allow_html=True)
st.write("---")

# Permissão de edição
senha = st.text_input("🔐 Digite a senha para adicionar ou editar presentes:", type="password")

pode_editar = senha == SENHA_CORRETA

if pode_editar:
    st.success("🔓 Edição Liberada!")

# Formulário para adicionar presente
if pode_editar:
    with st.form(key="form_presentes"):
        item = st.text_input("🎁 Nome do Presente", placeholder="Ex: Fritadeira Elétrica Mondial")
        link = st.text_input("🔗 Link do Produto", placeholder="Cole aqui o link do site")
        foto = st.file_uploader("📸 Enviar Foto do Produto (opcional)", type=["jpg", "jpeg", "png"])

        submit = st.form_submit_button("Adicionar Presente 🎀")

        if submit:
            if item and link:
                foto_path = ""
                if foto is not None:
                    foto_path = f"fotos_presentes/{foto.name}"
                    with open(foto_path, "wb") as f:
                        f.write(foto.read())

                presentes.append({
                    "item": item,
                    "status": "Disponível",
                    "comprador": "",
                    "link": link,
                    "foto": foto_path
                })
                salvar_presentes(presentes)
                st.success(f"✅ Presente '{item}' adicionado!")
                st.experimental_rerun()
            else:
                st.error("⚠️ Por favor, preencha o nome e o link do presente.")

st.write("---")

# Mostrar presentes
st.subheader("🎁 Presentes Disponíveis:")

if presentes:
    for i, presente in enumerate(presentes):
        col1, col2 = st.columns([6, 2])

        with col1:
            if presente["foto"]:
                try:
                    img = Image.open(presente["foto"])
                    st.image(img, width=150)
                except:
                    st.warning("Erro ao carregar imagem.")

            status_emoji = "✅" if presente["status"] == "Disponível" else "❌"
            st.markdown(f"**{status_emoji} {presente['item']}**")

            st.markdown(f"[🔗 Ver Produto]({presente['link']})")

            if presente["status"] == "Indisponível":
                st.info(f"🎉 Já comprado por: **{presente['comprador']}**")

        with col2:
            if presente["status"] == "Disponível":
                if st.button("Marcar como Comprado 🎁", key=f"comprar_{i}"):
                    comprador = st.text_input(f"Digite seu nome para confirmar:", key=f"input_{i}")
                    if comprador:
                        presentes[i]["status"] = "Indisponível"
                        presentes[i]["comprador"] = comprador
                        salvar_presentes(presentes)
                        st.success("🎉 Presente marcado como comprado!")
                        st.experimental_rerun()

            if pode_editar:
                if st.button("❌ Remover", key=f"remover_{i}"):
                    presentes.pop(i)
                    salvar_presentes(presentes)
                    st.success("Presente removido!")
                    st.experimental_rerun()
else:
    st.info("Nenhum presente adicionado ainda. 🎁")

st.write("---")

# Botão para gerar e baixar o Excel
if st.button("📥 Gerar e Baixar Lista em Excel"):
    if presentes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Presentes"

        headers = ["Item", "Status", "Nome de Quem Comprou", "Link do Produto"]
        ws.append(headers)

        for presente in presentes:
            ws.append([presente['item'], presente['status'], presente['comprador'], presente['link']])

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 30

        # Formatação condicional
        status_col = "B"
        formula_disponivel = f'${status_col}2="Disponível"'
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100")
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{len(presentes)+1}",
            FormulaRule(formula=[formula_disponivel], font=green_font, fill=green_fill)
        )

        formula_indisponivel = f'${status_col}2="Indisponível"'
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{len(presentes)+1}",
            FormulaRule(formula=[formula_indisponivel], font=red_fill, fill=red_fill)
        )

        # Salvar Excel temporário
        wb.save("Lista_Presentes_Casamento.xlsx")

        with open("Lista_Presentes_Casamento.xlsx", "rb") as f:
            st.download_button(
                label="📄 Baixar Planilha Excel",
                data=f,
                file_name="Lista_Presentes_Casamento.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("⚠️ Adicione ao menos um presente antes de gerar a planilha.")
